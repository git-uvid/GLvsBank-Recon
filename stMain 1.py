import streamlit as st
import pandas as pd
import numpy as np
import io
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from stCreatePivot import createBankPivot,createGLPivot,createDifferenceGrid
from stExportXl import write_reconciliation_summary_sheet,export_formatted_excel
from stBankGL import filter_type, matched_cal, Missing_el,get_comparison_key
from stOutstanding import dim_party, ost_bank, new_outstanding

# Initialize session state
if 'gl_data' not in st.session_state:
    st.session_state.gl_data = None
if 'bank_data' not in st.session_state:
    st.session_state.bank_data = None
if 'outstanding_check_data' not in st.session_state:
    st.session_state.outstanding_check_data = None
#if 'export_df' not in st.session_state: # To store the processed DFs for export
    #st.session_state.export_df = None
if 'reconciliation_excel_buffer' not in st.session_state: # To store the final Excel buffer for download
    st.session_state.reconciliation_excel_buffer = None

# App title
st.title("GL vs Bank Statement Reconciliation")

# File uploaders
st.sidebar.header("Upload Files")
gl_file = st.sidebar.file_uploader("Upload GL File (GLCategorized.xlsx)", type=["xlsx"])
bank_file = st.sidebar.file_uploader("Upload Bank File (BankCategorized.xlsx)", type=["xlsx"])

#method to set the bg based on the comment
def formatCommentCol(comment):
    if comment == "Transaction Matched":
        return 'background-color: green; color: white'
    elif comment == "Transaction Number not available in bank statement":
        return 'background-color: blue; color: white'
    elif comment == "Transaction number matched but the transacted amount is different":
        return 'background-color: yellow; color: Black'
    else:
        return 'background-color: red; color: white'

def process_files():
    if gl_file and bank_file:
        try:
            # Read input files
            st.session_state.gl_data = pd.read_excel(gl_file, sheet_name='GL Categorized', dtype=str)
            st.session_state.bank_data = pd.read_excel(bank_file, sheet_name='Categorized', dtype=str)
            # Read outstanding checks from GL file
            st.session_state.outstanding_check_data = pd.read_excel(gl_file, sheet_name='Outstanding Check Report',dtype=str)
            st.sidebar.success("Files uploaded and processed successfully!")
            return True
        except Exception as e:
            st.sidebar.error(f"Error processing files: {str(e)}")
            return False
    else:
        st.sidebar.warning("Please upload both files to proceed")
        return False
    
if st.sidebar.button("Process Files"):
    process_files()

def run_reconciliation():
     if st.session_state.gl_data is None or st.session_state.bank_data is None:
        st.warning("Please upload required files")
        return
     with st.spinner("Running reconciliation..."):
        try:
            # Select required columns and set data types
            gl_required_cols = st.session_state.gl_data[[
                'CO', 'AU', 'Acct', 'Sub Acct', 'Project', 'Period Name', 'Source', 'Category', 'Journal Name', 
                'Entered DR', 'Entered CR', 'Accounted DR', 'Accounted CR', 'Transaction Number', 'Transaction Date', 
                'Transaction Amount', 'Party Number', 'Party Name', 'Type', 'Accounted Sum'
            ]]

            bank_required_cols = st.session_state.bank_data[[
                'Bank reference', 'Customer reference', 'TRN TYPE', 'TRN status', 'Value date', 
                'Credit amount', 'Debit amount', 'Time', 'Post date'
            ]]

            gl_required_cols = gl_required_cols.astype({
                'CO': 'string', 'AU': 'string', 'Acct': 'string', 'Sub Acct': 'string', 'Project': 'string', 
                'Period Name': 'string', 'Source': 'string', 'Category': 'string', 'Journal Name': 'string', 
                'Entered DR': 'float', 'Entered CR': 'float', 'Accounted DR': 'float', 'Accounted CR': 'float',
                'Transaction Number': 'string', 'Transaction Date': 'string', 'Transaction Amount': 'float', 
                'Party Number': 'string', 'Party Name': 'string', 'Type': 'string', 'Accounted Sum': 'float'
            })

            bank_required_cols = bank_required_cols.astype({
                'Bank reference': 'string', 'Customer reference': 'string', 'TRN TYPE': 'string', 
                'TRN status': 'string', 'Value date': 'string', 'Credit amount': 'float', 
                'Debit amount': 'float', 'Time': 'string', 'Post date': 'string'
            })

            outstanding_check = st.session_state.outstanding_check_data.astype({
                'Check number': 'string', 'Date posted': 'string', 'Vendor Name': 'string',
                'Amount': 'float', 'Cleared?': 'string'
            })

            # Missing Transaction Number
            Missing_el(gl_required_cols,'Transaction Number','Tr')

            # Remove leading zeroes
            gl_required_cols['Transaction Number'] = gl_required_cols['Transaction Number'].str.lstrip('0')
            bank_required_cols['Bank reference'] = bank_required_cols['Bank reference'].str.lstrip('0')
            bank_required_cols['Customer reference'] = bank_required_cols['Customer reference'].str.lstrip('0')

            #fill missing ['Transaction Date','Transaction Amount', 'Party Number', 'Party Name'] with 'NA'
            gl_required_cols[['Transaction Date','Transaction Amount', 'Party Number', 'Party Name']] = gl_required_cols[['Transaction Date',
                                                                                                                           'Transaction Amount', 
                                                                                                                           'Party Number','Party Name']].fillna('NA')
# #---------------------------------------------Party dimension table---------------------------------------
            mrg_final_party_df = dim_party(gl_required_cols)    
#-------------------------------------------------------------------------Date posted table------------------------------------------------------
            #Dim table like df to get data posted in oustanding check report
            dateposted_req_cols = gl_required_cols[['Transaction Number','Transaction Date','Type']]
            dateposted_req_cols = dateposted_req_cols[dateposted_req_cols['Type'] == 'Checks']
            dateposted_req_cols = dateposted_req_cols[['Transaction Number','Transaction Date']]
            dateposted_req_cols = dateposted_req_cols.drop_duplicates()
#-------------------------------------------------------------------------------------------------------------------------------------------            
            # Aggregate GL data
            gl_agg = gl_required_cols.groupby([
                'CO', 'AU', 'Acct', 'Sub Acct', 'Project', 'Period Name', 'Source',
                 'Transaction Number', 'Type',
            ], as_index=False)['Accounted Sum'].sum()

            # Create comparison key for bank data
            bank_required_cols['comparsion_key'] = bank_required_cols.apply(get_comparison_key, axis=1)

            # Rename AR Module to AR
            def rename_ARModule(row):
                if row['TRN TYPE'] == "AR Module":
                    return 'AR'
                else:
                    return row['TRN TYPE']
                
            bank_required_cols['TRN TYPE'] = bank_required_cols.apply(rename_ARModule, axis=1)

            # Filter records
            Filter = ['Checks', 'Wires', 'AR']
            gl_req_cols_checks  = filter_type(gl_agg, 'Type',Filter)     
            bank_req_cols_checks = filter_type(bank_required_cols, 'TRN TYPE', Filter)

            # Merge GL and bank data
            matched_gl_bank = pd.merge(
                gl_req_cols_checks,
                bank_req_cols_checks,
                left_on=['Transaction Number', 'Type'],
                right_on=['comparsion_key', 'TRN TYPE'],
                how='outer'
            )

            matched_gl_bank = matched_cal(matched_gl_bank)


            # Format matched GL and bank data
            matched_gl_bank_formatted = matched_gl_bank.copy()
            matched_gl_bank_formatted.drop('Bank reference', axis=1, inplace=True)
            matched_gl_bank_formatted.drop('Customer reference', axis=1, inplace=True)

            matched_gl_bank_formatted.columns = [
                'Key_Transaction Number', 'GL_CO', 'GL_AU', 'GL_Acct', 'GL_Sub Acct', 'GL_Project', 
                'GL_Period Name', 'GL_Source','GL_Type', 'GL_Accounted Sum','Bnk_TRN TYPE', 'Bnk_TRN status', 
                'Bnk_Value date', 'Bnk_Credit amount', 'Bnk_Debit amount','Bnk_Accounted Sum','Bnk_Time', 'Bnk_Post date', 'Bnk_Comparsion_Key',
                'variance', 'comment'
            ]

#             #format gl bank matched excel
            matched_gl_bank_formatted = matched_gl_bank_formatted.style \
                 .map(formatCommentCol, subset=['comment']) \
                 .set_properties(**{
                                     'border': '1px solid black',
                                     'border-color': 'black'})
#--------------------------------------------------------Outstanding checks---------------------------------------------
            
            # # Process outstanding checks
            ost_bank_chks = ost_bank(outstanding_check, bank_req_cols_checks)

            # Add new outstanding checks from GL

            trans_not_inbank_reqcols_ost = new_outstanding(matched_gl_bank,ost_bank_chks)

            # Combine outstanding checks
            ost_bank_chks_final = pd.concat([ost_bank_chks, trans_not_inbank_reqcols_ost], ignore_index=True)
            ost_bank_chks_final[['Amount', 'Credit amount', 'Debit amount']] = ost_bank_chks_final[
                ['Amount', 'Credit amount', 'Debit amount']
            ].fillna(0)

            #get party name by merging with party dim df to get vendor name
            ost_bank_chks_final = pd.merge(ost_bank_chks_final,mrg_final_party_df,
                                           left_on='Check number',right_on='Transaction Number',    
                                           how='left')

            ost_bank_chks_final['Vendor Name'] = np.where(ost_bank_chks_final['Vendor Name'].isnull(),
                                                          ost_bank_chks_final['Party Name'],ost_bank_chks_final['Vendor Name'] )
            
            ost_bank_chks_final = ost_bank_chks_final[['Check number','Date posted','Vendor Name','Amount','Cleared?','Bank reference',
                                                       'Customer reference','TRN TYPE','TRN status','Value date','Credit amount',
                                                        'Debit amount','Time','Post date','comparsion_key','variance','updated status']]
            #get date posted by mergin with date dim df
            ost_bank_chks_final = pd.merge(ost_bank_chks_final,dateposted_req_cols,
                                           left_on='Check number',right_on='Transaction Number',
                                           how='left')
            ost_bank_chks_final['Date posted'] = np.where(ost_bank_chks_final['Date posted'].isnull(),
                                                          ost_bank_chks_final['Transaction Date'],ost_bank_chks_final['Date posted'])
            
            ost_bank_chks_final = ost_bank_chks_final[['Check number','Date posted','Vendor Name','Amount','Cleared?','Bank reference',
                                                       'Customer reference','TRN TYPE','TRN status','Value date','Credit amount',
                                                        'Debit amount','Time','Post date','comparsion_key','variance','updated status']]
            
            ost_bank_chks_formatStyle = ost_bank_chks_final .style \
                .set_properties(**{
                                    'border': '1px solid black',
                                    'border-color': 'black'})
#----------------------------------------------------Pivot section---------------------------------------------------------            
            bank_pivot = createBankPivot(bank_required_cols)
            gl_pivot = createGLPivot(gl_required_cols)
            diff_grid = createDifferenceGrid(bank_pivot,gl_pivot)
#---------------------------------------------------------------------------------------------------------------------------
#---------------------------------------- Orchestrate Excel Writing -------------------------------------------------------
            output = io.BytesIO()
            writer = pd.ExcelWriter(output, engine='xlsxwriter')
            # 1. Write the combined reconciliation summary sheet
            write_reconciliation_summary_sheet(
                writer,
                bank_pivot, # Pass raw pivot for this function's layout
                gl_pivot,   # Pass raw pivot for this function's layout
                diff_grid,
                sheet_name="pivot",
                header_bg_color='#4472C4', # Blue
                header_text_color="#FBEFEF",
                data_cell_border_color='gray',
                spacing_rows=2,
                spacing_cols=2
            )            
            dataframes_toexport = {'GLvsBank':matched_gl_bank_formatted,
                                   'OutstandingCheck':ost_bank_chks_formatStyle}
            # 3. Write other sheets using the same writer
            export_formatted_excel(
                dataframes_toexport,
                writer_obj=writer, # Pass the existing writer
                header_bg_color='#2F5496', # Blue for these sheets
                header_text_color='white'
            )

            writer.close()
            output.seek(0)
            st.success("Reconciliation completed and Excel report generated successfully!")
            return output

            return dataframes_toexport
        except Exception as e:
            st.error(f"Error during reconciliation: {str(e)}")
            return False
#-----------------------------------------Trigger Reconciliation--------------------------------------------------------------------------
if st.button("Run Reconciliation"):
    st.session_state.reconciliation_excel_buffer= run_reconciliation()

#-----------------------------------------Trigger Download--------------------------------------------------------------------------    
if st.session_state.reconciliation_excel_buffer:
    st.markdown("---")
    st.write("Reconciliation completed. Click the button below to generate and download the formatted Excel report.")
    st.download_button(
        label="Download Comprehensive Excel Report",
        data=st.session_state.reconciliation_excel_buffer,
        file_name="financial_reconciliation_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="Click to download the Excel file with all formatted reports."
    )
else:
    pass